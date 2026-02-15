"""
HTTP Router for Finance Module
Database: PostgreSQL (catering_db)
"""

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, desc, and_, extract
from sqlalchemy.orm import selectinload
from typing import List, Optional
from uuid import UUID
from datetime import datetime, date, timedelta
from decimal import Decimal
from pydantic import BaseModel
import logging

logger = logging.getLogger(__name__)

from backend.core.database import get_db
from backend.core.dependencies import get_current_tenant, CurrentTenant
from backend.modules.finance.domain.models import (
    AccountModel, JournalModel, JournalLineModel, FinanceTransactionModel,
    AccountingPeriodModel, BudgetModel, BudgetLineModel,
    PeriodAuditLogModel, PeriodCloseChecklistModel
)
from backend.modules.finance.domain.entities import Account, AccountBase, Journal, JournalBase
from backend.modules.order.domain.models import OrderModel, OrderPaymentModel

router = APIRouter(tags=["Finance Core"])




# ============ RESPONSE MODELS ============

class DashboardStats(BaseModel):
    revenue: dict
    expenses: dict
    profit: dict
    receivables: dict
    payables: dict

class TransactionResponse(BaseModel):
    id: UUID
    code: str
    type: str
    category: Optional[str]
    amount: Decimal
    payment_method: Optional[str]
    description: Optional[str]
    transaction_date: date
    reference_type: Optional[str]
    created_at: datetime
    
    class Config:
        from_attributes = True

class RecentTransaction(BaseModel):
    date: date
    description: str
    type: str  # THU or CHI
    amount: Decimal


# ============ DASHBOARD API ============

@router.get("/dashboard", response_model=DashboardStats)
async def get_dashboard_stats(tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """
    Get finance dashboard statistics.
    Integrates with Orders for revenue and receivables.
    """
    now = datetime.now()
    current_month_start = date(now.year, now.month, 1)
    
    # Previous month
    if now.month == 1:
        prev_month_start = date(now.year - 1, 12, 1)
        prev_month_end = date(now.year - 1, 12, 31)
    else:
        prev_month_start = date(now.year, now.month - 1, 1)
        prev_month_end = current_month_start - timedelta(days=1)
    
    # === REVENUE FROM ORDERS ===
    # Current month revenue (from paid orders)
    current_revenue_result = await db.execute(
        select(func.coalesce(func.sum(OrderModel.paid_amount), 0))
        .where(
            OrderModel.tenant_id == tenant_id,
            OrderModel.status.in_(['COMPLETED', 'PAID']),
            func.date(OrderModel.created_at) >= current_month_start
        )
    )
    current_revenue = current_revenue_result.scalar() or Decimal(0)
    
    # Previous month revenue
    prev_revenue_result = await db.execute(
        select(func.coalesce(func.sum(OrderModel.paid_amount), 0))
        .where(
            OrderModel.tenant_id == tenant_id,
            OrderModel.status.in_(['COMPLETED', 'PAID']),
            func.date(OrderModel.created_at) >= prev_month_start,
            func.date(OrderModel.created_at) <= prev_month_end
        )
    )
    prev_revenue = prev_revenue_result.scalar() or Decimal(0)
    
    # Revenue growth
    if prev_revenue > 0:
        revenue_growth = ((current_revenue - prev_revenue) / prev_revenue * 100)
    else:
        revenue_growth = Decimal(100) if current_revenue > 0 else Decimal(0)
    
    # === EXPENSES (from finance_transactions) ===
    current_expenses_result = await db.execute(
        select(func.coalesce(func.sum(FinanceTransactionModel.amount), 0))
        .where(
            FinanceTransactionModel.tenant_id == tenant_id,
            FinanceTransactionModel.type == 'PAYMENT',
            FinanceTransactionModel.transaction_date >= current_month_start
        )
    )
    current_expenses = current_expenses_result.scalar() or Decimal(0)
    
    prev_expenses_result = await db.execute(
        select(func.coalesce(func.sum(FinanceTransactionModel.amount), 0))
        .where(
            FinanceTransactionModel.tenant_id == tenant_id,
            FinanceTransactionModel.type == 'PAYMENT',
            FinanceTransactionModel.transaction_date >= prev_month_start,
            FinanceTransactionModel.transaction_date <= prev_month_end
        )
    )
    prev_expenses = prev_expenses_result.scalar() or Decimal(0)
    
    if prev_expenses > 0:
        expenses_growth = ((current_expenses - prev_expenses) / prev_expenses * 100)
    else:
        expenses_growth = Decimal(0)
    
    # === PROFIT ===
    current_profit = current_revenue - current_expenses
    if current_revenue > 0:
        margin_percent = (current_profit / current_revenue * 100)
    else:
        margin_percent = Decimal(0)
    
    # === RECEIVABLES (unpaid order amounts) ===
    try:
        receivables_total_result = await db.execute(
            select(func.coalesce(func.sum(OrderModel.balance_amount), 0))
            .where(
                OrderModel.tenant_id == tenant_id,
                OrderModel.balance_amount > 0,
                OrderModel.status.notin_(['CANCELLED'])
            )
        )
        receivables_total = receivables_total_result.scalar() or Decimal(0)
        receivables_overdue = Decimal(0)  # Simplified for now
    except Exception:
        receivables_total = Decimal(0)
        receivables_overdue = Decimal(0)
    
    # === PAYABLES (from Procurement POs) ===
    try:
        from backend.modules.procurement.domain.models import PurchaseOrderModel as POModel
        
        # Total unpaid POs
        payables_result = await db.execute(
            select(func.coalesce(func.sum(POModel.total_amount - func.coalesce(POModel.paid_amount, 0)), 0))
            .where(
                POModel.tenant_id == tenant_id,
                POModel.status.notin_(['PAID', 'CANCELLED', 'DRAFT'])
            )
        )
        payables_total = payables_result.scalar() or Decimal(0)
        
        # Due soon (within 7 days)
        from datetime import timezone as tz
        now_utc = datetime.now(tz.utc)
        due_soon_result = await db.execute(
            select(func.coalesce(func.sum(POModel.total_amount - func.coalesce(POModel.paid_amount, 0)), 0))
            .where(
                POModel.tenant_id == tenant_id,
                POModel.status.notin_(['PAID', 'CANCELLED', 'DRAFT']),
                POModel.due_date <= now_utc + timedelta(days=7)
            )
        )
        payables_due_soon = due_soon_result.scalar() or Decimal(0)
    except Exception:
        payables_total = Decimal(0)
        payables_due_soon = Decimal(0)
    
    return DashboardStats(
        revenue={
            "current_month": float(current_revenue),
            "previous_month": float(prev_revenue),
            "growth_percent": round(float(revenue_growth), 2)
        },
        expenses={
            "current_month": float(current_expenses),
            "previous_month": float(prev_expenses),
            "growth_percent": round(float(expenses_growth), 2)
        },
        profit={
            "current_month": float(current_profit),
            "margin_percent": round(float(margin_percent), 2)
        },
        receivables={
            "total": float(receivables_total),
            "overdue": float(receivables_overdue)
        },
        payables={
            "total": float(payables_total),
            "due_soon": float(payables_due_soon)
        }
    )


@router.get("/recent-transactions", response_model=List[RecentTransaction])
async def get_recent_transactions(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    limit: int = Query(10, le=50)
):
    """
    Get recent transactions combining Order Payments and Finance Transactions.
    """
    transactions = []
    
    # Get recent order payments
    payments_result = await db.execute(
        select(OrderPaymentModel)
        .join(OrderModel)
        .where(OrderModel.tenant_id == tenant_id)
        .order_by(desc(OrderPaymentModel.payment_date))
        .limit(limit)
    )
    payments = payments_result.scalars().all()
    
    for p in payments:
        transactions.append(RecentTransaction(
            date=p.payment_date.date() if p.payment_date else date.today(),
            description=f"Thu tiền đơn hàng - {p.note or 'Thanh toán'}",
            type="THU",
            amount=p.amount or Decimal(0)
        ))
    
    # Get recent finance transactions
    finance_result = await db.execute(
        select(FinanceTransactionModel)
        .where(FinanceTransactionModel.tenant_id == tenant_id)
        .order_by(desc(FinanceTransactionModel.created_at))
        .limit(limit)
    )
    finance_trans = finance_result.scalars().all()
    
    for t in finance_trans:
        transactions.append(RecentTransaction(
            date=t.transaction_date,
            description=t.description or f"Giao dịch {t.code}",
            type="THU" if t.type == "RECEIPT" else "CHI",
            amount=t.amount
        ))
    
    # Sort by date desc and limit
    transactions.sort(key=lambda x: x.date, reverse=True)
    return transactions[:limit]


# ============ STATS FOR CHARTS ============

@router.get("/stats/monthly")
async def get_monthly_stats(
    tenant_id: UUID = Depends(get_current_tenant), 
    db: AsyncSession = Depends(get_db),
    months: int = Query(12, le=24),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
):
    """
    Get monthly revenue/expense stats for charts.
    Returns data for Xu Hướng Thu Chi chart.
    Supports date range filtering via start_date/end_date params.
    """
    now = datetime.now()
    
    # Parse date range if provided
    filter_start = None
    filter_end = None
    if start_date:
        try:
            filter_start = date.fromisoformat(start_date)
        except ValueError:
            pass
    if end_date:
        try:
            filter_end = date.fromisoformat(end_date)
        except ValueError:
            pass
    
    # If date range provided, calculate months to cover
    if filter_start and filter_end:
        # Calculate number of months between start and end
        month_diff = (filter_end.year - filter_start.year) * 12 + (filter_end.month - filter_start.month) + 1
        months = max(month_diff, 1)
    
    result = []
    
    for i in range(months - 1, -1, -1):
        # Calculate month
        if filter_start and filter_end:
            # Use filter_start as base, iterate forward
            month_offset = months - 1 - i
            year = filter_start.year + (filter_start.month + month_offset - 1) // 12
            month_num = (filter_start.month + month_offset - 1) % 12 + 1
            month_start = date(year, month_num, 1)
        else:
            month_date = now - timedelta(days=i * 30)
            month_start = date(month_date.year, month_date.month, 1)
        
        if month_start.month == 12:
            month_end = date(month_start.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)
        
        # Revenue from Order Payments (Thu)
        receipts_result = await db.execute(
            select(func.coalesce(func.sum(OrderPaymentModel.amount), 0))
            .join(OrderModel, OrderPaymentModel.order_id == OrderModel.id)
            .where(
                OrderModel.tenant_id == tenant_id,
                func.date(OrderPaymentModel.payment_date) >= month_start,
                func.date(OrderPaymentModel.payment_date) <= month_end
            )
        )
        thu = float(receipts_result.scalar() or 0)
        
        # Expenses from Finance Transactions (Chi)
        payments_result = await db.execute(
            select(func.coalesce(func.sum(FinanceTransactionModel.amount), 0))
            .where(
                FinanceTransactionModel.tenant_id == tenant_id,
                FinanceTransactionModel.type == 'PAYMENT',
                FinanceTransactionModel.transaction_date >= month_start,
                FinanceTransactionModel.transaction_date <= month_end
            )
        )
        chi = float(payments_result.scalar() or 0)
        
        result.append({
            "month": month_start.strftime('T%m'),  # T01, T02, ...
            "period": month_start.strftime('%m/%Y'),
            "thu": thu,
            "chi": chi
        })
    
    return result


@router.get("/stats/expenses-by-category")
async def get_expenses_by_category(
    tenant_id: UUID = Depends(get_current_tenant), 
    db: AsyncSession = Depends(get_db),
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
):
    """
    Get expenses grouped by category for pie chart.
    Returns data for Phân Loại Chi Phí chart.
    Supports date range filtering via start_date/end_date params.
    """
    # Parse date range
    if start_date:
        try:
            filter_start = date.fromisoformat(start_date)
        except ValueError:
            filter_start = date(datetime.now().year, datetime.now().month, 1)
    else:
        filter_start = date(datetime.now().year, datetime.now().month, 1)
    
    if end_date:
        try:
            filter_end = date.fromisoformat(end_date)
        except ValueError:
            filter_end = date.today()
    else:
        filter_end = date.today()
    
    # Get expenses by category for the date range
    result = await db.execute(
        select(
            FinanceTransactionModel.category,
            func.sum(FinanceTransactionModel.amount).label('total')
        )
        .where(
            FinanceTransactionModel.tenant_id == tenant_id,
            FinanceTransactionModel.type == 'PAYMENT',
            FinanceTransactionModel.transaction_date >= filter_start,
            FinanceTransactionModel.transaction_date <= filter_end
        )
        .group_by(FinanceTransactionModel.category)
    )
    rows = result.all()
    
    # Map to Vietnamese category names
    category_map = {
        'NGUYENLIEU': 'Nguyên liệu',
        'NHANCONG': 'Nhân công',
        'VANHANH': 'Vận hành',
        'MARKETING': 'Marketing',
        'KHAC': 'Khác',
        None: 'Khác'
    }
    
    categories = []
    for row in rows:
        cat_code = row[0]
        amount = float(row[1] or 0)
        categories.append({
            "category": category_map.get(cat_code, cat_code or 'Khác'),
            "amount": amount
        })
    
    # If no data, return mock categories with 0
    if not categories:
        categories = [
            {"category": "Nguyên liệu", "amount": 0},
            {"category": "Nhân công", "amount": 0},
            {"category": "Vận hành", "amount": 0},
            {"category": "Marketing", "amount": 0},
            {"category": "Khác", "amount": 0}
        ]
    
    return categories


# ============ ACCOUNTS (Chart of Accounts) ============

@router.get("/accounts", response_model=List[Account])
async def list_accounts(tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """Get Chart of Accounts from PostgreSQL"""
    result = await db.execute(
        select(AccountModel)
        .where(AccountModel.tenant_id == tenant_id)
        .order_by(AccountModel.code)
    )
    accounts = result.scalars().all()
    return accounts


@router.post("/accounts", response_model=Account)
async def create_account(data: AccountBase, tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """Create a new account in Chart of Accounts"""
    # Check if code already exists
    existing = await db.execute(
        select(AccountModel).where(
            AccountModel.tenant_id == tenant_id,
            AccountModel.code == data.code
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"Mã tài khoản {data.code} đã tồn tại")
    
    new_account = AccountModel(
        tenant_id=tenant_id,
        code=data.code,
        name=data.name,
        type=data.type,
        is_active=data.is_active
    )
    db.add(new_account)
    await db.commit()
    await db.refresh(new_account)
    return new_account


# ============ TRANSACTIONS ============

class TransactionCreate(BaseModel):
    type: str  # RECEIPT or PAYMENT
    category: Optional[str] = None
    amount: Decimal
    payment_method: str = "CASH"
    description: Optional[str] = None
    transaction_date: date
    reference_id: Optional[UUID] = None
    reference_type: Optional[str] = None


@router.get("/transactions", response_model=List[TransactionResponse])
async def list_transactions(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    type: Optional[str] = Query(None, description="RECEIPT or PAYMENT"),
    skip: int = 0,
    limit: int = 50
):
    """List finance transactions"""
    query = select(FinanceTransactionModel).where(
        FinanceTransactionModel.tenant_id == tenant_id
    )
    
    if type:
        query = query.where(FinanceTransactionModel.type == type)
    
    query = query.order_by(desc(FinanceTransactionModel.transaction_date)).offset(skip).limit(limit)
    
    result = await db.execute(query)
    return result.scalars().all()


@router.post("/transactions", response_model=TransactionResponse)
async def create_transaction(data: TransactionCreate, tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """Create a finance transaction (Thu/Chi)"""
    import random
    
    # Generate transaction code
    now = datetime.now()
    prefix = "THU" if data.type == "RECEIPT" else "CHI"
    random_suffix = random.randint(1000, 9999)
    code = f"{prefix}-{now.strftime('%Y%m')}-{random_suffix}"
    
    new_transaction = FinanceTransactionModel(
        tenant_id=tenant_id,
        code=code,
        type=data.type,
        category=data.category,
        amount=data.amount,
        payment_method=data.payment_method,
        description=data.description,
        transaction_date=data.transaction_date,
        reference_id=data.reference_id,
        reference_type=data.reference_type
    )
    
    db.add(new_transaction)
    
    # R1: Update order.expenses_amount if expense is linked to an order
    if data.reference_type == "ORDER" and data.reference_id and data.type == "PAYMENT":
        order_result = await db.execute(
            select(OrderModel).where(
                OrderModel.id == data.reference_id,
                OrderModel.tenant_id == tenant_id
            )
        )
        order = order_result.scalar_one_or_none()
        if order:
            current_expenses = order.expenses_amount or Decimal(0)
            order.expenses_amount = current_expenses + data.amount
            logger.info(f"R1: Updated order {order.code} expenses_amount to {order.expenses_amount}")
    
    await db.commit()
    await db.refresh(new_transaction)
    return new_transaction


# ============ RECEIVABLES (from Orders) ============

class ReceivableItem(BaseModel):
    order_id: UUID
    order_code: str
    customer_name: str
    customer_phone: Optional[str]
    final_amount: Decimal
    paid_amount: Decimal
    balance_amount: Decimal
    event_date: Optional[datetime]
    status: str
    aging_bucket: str


@router.get("/receivables", response_model=List[ReceivableItem])
async def list_receivables(tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """
    List accounts receivable (unpaid order amounts).
    Integrated with Order module.
    """
    now = datetime.now()
    
    result = await db.execute(
        select(OrderModel)
        .where(
            OrderModel.tenant_id == tenant_id,
            OrderModel.balance_amount > 0,
            OrderModel.status.notin_(['CANCELLED'])
        )
        .order_by(OrderModel.event_date)
    )
    orders = result.scalars().all()
    
    receivables = []
    for o in orders:
        # Calculate aging bucket
        if o.event_date:
            days_old = (now - o.event_date).days if o.event_date < now else 0
            if days_old > 90:
                bucket = "90+"
            elif days_old > 60:
                bucket = "60-90"
            elif days_old > 30:
                bucket = "30-60"
            else:
                bucket = "CURRENT"
        else:
            bucket = "CURRENT"
        
        receivables.append(ReceivableItem(
            order_id=o.id,
            order_code=o.code,
            customer_name=o.customer_name or "N/A",
            customer_phone=o.customer_phone,
            final_amount=o.final_amount or Decimal(0),
            paid_amount=o.paid_amount or Decimal(0),
            balance_amount=o.balance_amount or Decimal(0),
            event_date=o.event_date,
            status=o.status,
            aging_bucket=bucket
        ))
    
    return receivables


# ============ R2: RECEIVABLES ALERTS ============

class ReceivableAlertItem(BaseModel):
    """Alert item for overdue receivables"""
    order_id: UUID
    order_code: str
    customer_name: str
    customer_phone: Optional[str]
    balance_amount: Decimal
    event_date: Optional[datetime]
    days_overdue: int
    priority: str  # HIGH, MEDIUM, LOW


class ReceivableAlertsSummary(BaseModel):
    total_overdue: int
    total_amount: Decimal
    high_priority_count: int
    alerts: List[ReceivableAlertItem]


@router.get("/receivables/alerts", response_model=ReceivableAlertsSummary)
async def get_receivables_alerts(
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
    days_threshold: int = Query(7, description="Days overdue threshold")
):
    """
    R2: Get overdue receivables alerts.
    Returns orders with balance > 0 and event_date older than threshold days.
    """
    now = datetime.now()
    threshold_date = now - timedelta(days=days_threshold)
    
    result = await db.execute(
        select(OrderModel)
        .where(
            OrderModel.tenant_id == tenant_id,
            OrderModel.balance_amount > 0,
            OrderModel.status.notin_(['CANCELLED']),
            OrderModel.event_date < threshold_date
        )
        .order_by(OrderModel.event_date.asc())
    )
    orders = result.scalars().all()
    
    alerts = []
    total_amount = Decimal(0)
    high_priority_count = 0
    
    for o in orders:
        days_overdue = (now - o.event_date).days if o.event_date else 0
        
        # Priority: HIGH (>14 days), MEDIUM (7-14), LOW (<7)
        if days_overdue > 14:
            priority = "HIGH"
            high_priority_count += 1
        elif days_overdue > 7:
            priority = "MEDIUM"
        else:
            priority = "LOW"
        
        balance = o.balance_amount or Decimal(0)
        total_amount += balance
        
        alerts.append(ReceivableAlertItem(
            order_id=o.id,
            order_code=o.code,
            customer_name=o.customer_name or "N/A",
            customer_phone=o.customer_phone,
            balance_amount=balance,
            event_date=o.event_date,
            days_overdue=days_overdue,
            priority=priority
        ))
    
    return ReceivableAlertsSummary(
        total_overdue=len(alerts),
        total_amount=total_amount,
        high_priority_count=high_priority_count,
        alerts=alerts
    )


# ============ BULK PAYMENTS ============

class BulkPaymentItem(BaseModel):
    """Single payment in bulk request"""
    order_id: UUID
    amount: Decimal
    payment_method: str = "CASH"
    payment_date: Optional[str] = None

class BulkPaymentRequest(BaseModel):
    """Request for bulk payments"""
    payments: List[BulkPaymentItem]

class BulkPaymentResult(BaseModel):
    """Result of single payment in bulk"""
    order_id: UUID
    order_code: str
    amount: Decimal
    success: bool
    error: Optional[str] = None

class BulkPaymentResponse(BaseModel):
    """Response for bulk payments"""
    total_processed: int
    success_count: int
    failed_count: int
    total_amount: Decimal
    results: List[BulkPaymentResult]

@router.post("/payments/bulk", response_model=BulkPaymentResponse)
async def record_bulk_payments(
    request: BulkPaymentRequest,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Record multiple payments at once.
    Used for bulk payment recording from receivables table.
    """
    from datetime import timezone as tz
    
    results = []
    success_count = 0
    failed_count = 0
    total_amount = Decimal(0)
    
    for payment in request.payments:
        try:
            # Get order
            order_result = await db.execute(
                select(OrderModel).where(
                    (OrderModel.id == payment.order_id) & 
                    (OrderModel.tenant_id == tenant_id)
                )
            )
            order = order_result.scalar_one_or_none()
            
            if not order:
                results.append(BulkPaymentResult(
                    order_id=payment.order_id,
                    order_code="N/A",
                    amount=payment.amount,
                    success=False,
                    error="Order not found"
                ))
                failed_count += 1
                continue
            
            # Validate payment amount
            if payment.amount <= 0:
                results.append(BulkPaymentResult(
                    order_id=payment.order_id,
                    order_code=order.code,
                    amount=payment.amount,
                    success=False,
                    error="Invalid payment amount"
                ))
                failed_count += 1
                continue
            
            # Create payment record
            payment_date = datetime.now(tz.utc)
            if payment.payment_date:
                try:
                    payment_date = datetime.strptime(payment.payment_date, "%Y-%m-%d").replace(tzinfo=tz.utc)
                except ValueError:
                    pass
            
            new_payment = OrderPaymentModel(
                tenant_id=tenant_id,
                order_id=payment.order_id,
                amount=payment.amount,
                payment_method=payment.payment_method,
                payment_date=payment_date
            )
            db.add(new_payment)
            
            # Update order totals
            order.paid_amount = (order.paid_amount or Decimal(0)) + payment.amount
            order.balance_amount = (order.final_amount or Decimal(0)) - order.paid_amount
            
            # Auto-transition to PAID if fully paid
            if order.balance_amount <= 0:
                order.status = 'PAID'
            
            order.updated_at = datetime.now(tz.utc)
            
            results.append(BulkPaymentResult(
                order_id=payment.order_id,
                order_code=order.code,
                amount=payment.amount,
                success=True
            ))
            success_count += 1
            total_amount += payment.amount
            
        except Exception as e:
            logger.error(f"Bulk payment error for order {payment.order_id}: {e}")
            results.append(BulkPaymentResult(
                order_id=payment.order_id,
                order_code="N/A",
                amount=payment.amount,
                success=False,
                error=str(e)
            ))
            failed_count += 1
    
    # Commit all successful payments
    await db.commit()
    
    return BulkPaymentResponse(
        total_processed=len(request.payments),
        success_count=success_count,
        failed_count=failed_count,
        total_amount=total_amount,
        results=results
    )


# ============ JOURNALS ============

@router.get("/journals")
async def list_journals(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    skip: int = 0,
    limit: int = 50
):
    """List journal entries"""
    result = await db.execute(
        select(JournalModel)
        .where(JournalModel.tenant_id == tenant_id)
        .options(selectinload(JournalModel.lines))
        .order_by(desc(JournalModel.date))
        .offset(skip)
        .limit(limit)
    )
    return result.scalars().all()


@router.post("/journals", response_model=Journal)
async def create_journal(data: JournalBase, tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """Create a journal entry with lines"""
    import random
    
    now = datetime.now()
    random_suffix = random.randint(100, 999)
    code = f"JNL-{now.strftime('%Y%m')}-{random_suffix}"
    
    # Calculate total (sum of debits)
    total = sum(line.debit for line in data.lines)
    
    new_journal = JournalModel(
        tenant_id=tenant_id,
        code=code,
        date=data.date,
        description=data.description,
        total_amount=total
    )
    db.add(new_journal)
    await db.flush()
    
    # Add lines
    for line in data.lines:
        new_line = JournalLineModel(
            tenant_id=tenant_id,
            journal_id=new_journal.id,
            account_id=line.account_id,
            debit=line.debit,
            credit=line.credit,
            description=line.description
        )
        db.add(new_line)
    
    await db.commit()
    await db.refresh(new_journal)
    return new_journal


@router.get("/journals/{journal_id}")
async def get_journal(
    journal_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Get a single journal entry with lines"""
    result = await db.execute(
        select(JournalModel)
        .where(JournalModel.id == journal_id, JournalModel.tenant_id == tenant_id)
        .options(selectinload(JournalModel.lines))
    )
    journal = result.scalar_one_or_none()
    
    if not journal:
        raise HTTPException(status_code=404, detail="Journal not found")
    
    return journal


@router.post("/journals/{journal_id}/post")
async def post_journal(
    journal_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Post a journal entry (FIN-001).
    Changes status from DRAFT to POSTED.
    Posted journals cannot be edited, only reversed.
    """
    result = await db.execute(
        select(JournalModel)
        .where(JournalModel.id == journal_id, JournalModel.tenant_id == tenant_id)
        .options(selectinload(JournalModel.lines))
    )
    journal = result.scalar_one_or_none()
    
    if not journal:
        raise HTTPException(status_code=404, detail="Journal not found")
    
    if journal.status == 'POSTED':
        raise HTTPException(status_code=400, detail="Journal is already posted")
    
    if journal.status == 'REVERSED':
        raise HTTPException(status_code=400, detail="Cannot post a reversed journal")
    
    # Validate balance: sum(debit) must equal sum(credit)
    total_debit = sum(line.debit or Decimal(0) for line in journal.lines)
    total_credit = sum(line.credit or Decimal(0) for line in journal.lines)
    
    if total_debit != total_credit:
        raise HTTPException(
            status_code=400, 
            detail=f"Journal is not balanced. Debit: {total_debit}, Credit: {total_credit}"
        )
    
    # Post the journal
    journal.status = 'POSTED'
    journal.posted_at = datetime.now()
    # journal.posted_by = current_user.id  # TODO: Add when auth is implemented
    
    await db.commit()
    
    return {
        "message": "Journal posted successfully",
        "journal_id": str(journal_id),
        "status": "POSTED",
        "posted_at": journal.posted_at.isoformat()
    }


@router.post("/journals/{journal_id}/reverse")
async def reverse_journal(
    journal_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Reverse a posted journal entry (FIN-001).
    Creates a new journal with opposite debit/credit amounts.
    Marks original journal as REVERSED.
    """
    result = await db.execute(
        select(JournalModel)
        .where(JournalModel.id == journal_id, JournalModel.tenant_id == tenant_id)
        .options(selectinload(JournalModel.lines))
    )
    journal = result.scalar_one_or_none()
    
    if not journal:
        raise HTTPException(status_code=404, detail="Journal not found")
    
    if journal.status != 'POSTED':
        raise HTTPException(status_code=400, detail="Only posted journals can be reversed")
    
    if journal.reversed_journal_id:
        raise HTTPException(status_code=400, detail="Journal has already been reversed")
    
    # Generate reversal journal code
    import random
    now = datetime.now()
    random_suffix = random.randint(100, 999)
    reversal_code = f"REV-{now.strftime('%Y%m')}-{random_suffix}"
    
    # Create reversal journal
    reversal_journal = JournalModel(
        tenant_id=tenant_id,
        code=reversal_code,
        date=now,
        description=f"Đảo ngược bút toán {journal.code}",
        total_amount=journal.total_amount,
        reference_id=journal.reference_id,
        reference_type=journal.reference_type,
        status='POSTED',
        posted_at=now
    )
    db.add(reversal_journal)
    await db.flush()
    
    # Create reversal lines (swap debit/credit)
    for line in journal.lines:
        reversal_line = JournalLineModel(
            tenant_id=tenant_id,
            journal_id=reversal_journal.id,
            account_id=line.account_id,
            debit=line.credit,  # Swap: original credit becomes debit
            credit=line.debit,  # Swap: original debit becomes credit
            description=f"Đảo: {line.description or ''}"
        )
        db.add(reversal_line)
    
    # Mark original as reversed
    journal.status = 'REVERSED'
    journal.reversed_journal_id = reversal_journal.id
    
    await db.commit()
    
    return {
        "message": "Journal reversed successfully",
        "original_journal_id": str(journal_id),
        "reversal_journal_id": str(reversal_journal.id),
        "reversal_code": reversal_code
    }


# ============ BALANCE SHEET REPORT (FIN-003) ============

class BalanceSheetItem(BaseModel):
    account_code: str
    account_name: str
    account_type: str
    debit_balance: float
    credit_balance: float
    net_balance: float


@router.get("/reports/balance-sheet")
async def get_balance_sheet_report(
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
    as_of_date: Optional[str] = Query(None, description="As of date (YYYY-MM-DD), defaults to today")
):
    """
    Balance Sheet Report (FIN-003).
    Assets = Liabilities + Equity
    Shows balances for all accounts as of a specific date.
    """
    if as_of_date:
        report_date = date.fromisoformat(as_of_date)
    else:
        report_date = date.today()
    
    # Get all accounts
    accounts_result = await db.execute(
        select(AccountModel)
        .where(AccountModel.tenant_id == tenant_id, AccountModel.is_active == True)
        .order_by(AccountModel.code)
    )
    accounts = accounts_result.scalars().all()
    
    account_balances = {}
    
    # Calculate balances from posted journals
    for account in accounts:
        # Sum all debits and credits from posted journals for this account
        balance_result = await db.execute(
            select(
                func.coalesce(func.sum(JournalLineModel.debit), 0).label('total_debit'),
                func.coalesce(func.sum(JournalLineModel.credit), 0).label('total_credit')
            )
            .join(JournalModel, JournalLineModel.journal_id == JournalModel.id)
            .where(
                JournalLineModel.tenant_id == tenant_id,
                JournalLineModel.account_id == account.id,
                JournalModel.status == 'POSTED',
                func.date(JournalModel.date) <= report_date
            )
        )
        row = balance_result.first()
        total_debit = float(row.total_debit or 0)
        total_credit = float(row.total_credit or 0)
        
        # Net balance depends on account type:
        # ASSET, EXPENSE: Debit is positive (debit - credit)
        # LIABILITY, EQUITY, REVENUE: Credit is positive (credit - debit)
        if account.type in ('ASSET', 'EXPENSE'):
            net_balance = total_debit - total_credit
        else:
            net_balance = total_credit - total_debit
        
        account_balances[account.id] = BalanceSheetItem(
            account_code=account.code,
            account_name=account.name,
            account_type=account.type,
            debit_balance=total_debit,
            credit_balance=total_credit,
            net_balance=net_balance
        )
    
    # Group by type
    assets = [b for b in account_balances.values() if b.account_type == 'ASSET' and b.net_balance != 0]
    liabilities = [b for b in account_balances.values() if b.account_type == 'LIABILITY' and b.net_balance != 0]
    equity = [b for b in account_balances.values() if b.account_type == 'EQUITY' and b.net_balance != 0]
    
    # Calculate totals
    total_assets = sum(b.net_balance for b in assets)
    total_liabilities = sum(b.net_balance for b in liabilities)
    total_equity = sum(b.net_balance for b in equity)
    
    # Add retained earnings (Revenue - Expense) to equity
    revenues = [b for b in account_balances.values() if b.account_type == 'REVENUE']
    expenses = [b for b in account_balances.values() if b.account_type == 'EXPENSE']
    retained_earnings = sum(b.net_balance for b in revenues) - sum(abs(b.net_balance) for b in expenses)
    
    return {
        "as_of_date": report_date.isoformat(),
        "sections": {
            "assets": {
                "items": [b.model_dump() for b in assets],
                "total": total_assets
            },
            "liabilities": {
                "items": [b.model_dump() for b in liabilities],
                "total": total_liabilities
            },
            "equity": {
                "items": [b.model_dump() for b in equity],
                "retained_earnings": retained_earnings,
                "total": total_equity + retained_earnings
            }
        },
        "summary": {
            "total_assets": total_assets,
            "total_liabilities": total_liabilities,
            "total_equity": total_equity + retained_earnings,
            "is_balanced": abs(total_assets - (total_liabilities + total_equity + retained_earnings)) < 0.01
        }
    }


# ============ PAYABLES (from Procurement/Purchase Orders) ============

from backend.modules.procurement.domain.models import PurchaseOrderModel, SupplierModel

class PayableItem(BaseModel):
    po_id: UUID
    po_code: str
    supplier_name: str
    supplier_phone: Optional[str]
    total_amount: Decimal
    paid_amount: Decimal = Decimal(0)  # Track payments made
    balance: Decimal  # total_amount - paid_amount
    status: str
    expected_delivery: Optional[datetime]
    created_at: datetime
    payment_terms_days: int = 30  # Days until due
    days_outstanding: int


@router.get("/payables", response_model=List[PayableItem])
async def list_payables(
    include_paid: bool = False,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    List accounts payable (purchase orders).
    By default only shows unpaid POs. Set include_paid=true to include paid POs.
    """
    from datetime import timezone
    now = datetime.now(timezone.utc)
    
    # Build status exclusion list
    excluded_statuses = ['CANCELLED', 'DRAFT']
    if not include_paid:
        excluded_statuses.append('PAID')
    
    result = await db.execute(
        select(PurchaseOrderModel)
        .options(selectinload(PurchaseOrderModel.supplier))
        .where(
            PurchaseOrderModel.tenant_id == tenant_id,
            PurchaseOrderModel.status.notin_(excluded_statuses)
        )
        .order_by(PurchaseOrderModel.created_at.desc())
    )
    orders = result.scalars().all()
    
    payables = []
    for o in orders:
        # Handle timezone-aware datetime comparison
        created = o.created_at
        if created:
            if created.tzinfo is None:
                created = created.replace(tzinfo=timezone.utc)
            days = (now - created).days
        else:
            days = 0
            
        total = o.total_amount or Decimal(0)
        paid = o.paid_amount or Decimal(0)
        balance = total - paid
        
        # Calculate payment terms in days
        term_days_map = {
            'IMMEDIATE': 0,
            'NET15': 15,
            'NET30': 30,
            'NET60': 60,
            'NET90': 90
        }
        payment_terms_days = term_days_map.get(o.payment_terms or 'NET30', 30)
        
        payables.append(PayableItem(
            po_id=o.id,
            po_code=o.code,
            supplier_name=o.supplier.name if o.supplier else "N/A",
            supplier_phone=o.supplier.phone if o.supplier else None,
            total_amount=total,
            paid_amount=paid,
            balance=balance,
            status=o.status,
            expected_delivery=o.expected_delivery,
            created_at=o.created_at,
            payment_terms_days=payment_terms_days,
            days_outstanding=days
        ))
    
    return payables


@router.get("/payables/summary")
async def get_payables_summary(tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """Get payables summary for dashboard"""
    # Total payables (not PAID)
    total_result = await db.execute(
        select(func.coalesce(func.sum(PurchaseOrderModel.total_amount), 0))
        .where(
            PurchaseOrderModel.tenant_id == tenant_id,
            PurchaseOrderModel.status.notin_(['PAID', 'CANCELLED', 'DRAFT'])
        )
    )
    total = total_result.scalar() or Decimal(0)
    
    # Count
    count_result = await db.execute(
        select(func.count(PurchaseOrderModel.id))
        .where(
            PurchaseOrderModel.tenant_id == tenant_id,
            PurchaseOrderModel.status.notin_(['PAID', 'CANCELLED', 'DRAFT'])
        )
    )
    count = count_result.scalar() or 0
    
    return {
        "total": float(total),
        "count": count
    }


# ============ PHASE 5B: PAYMENT SCHEDULE ============

class PaymentScheduleItem(BaseModel):
    po_id: UUID
    po_code: str
    supplier_name: str
    amount: float
    paid_amount: float
    balance: float
    payment_terms: str
    due_date: Optional[datetime]
    days_until_due: int
    status: str  # ON_TIME, DUE_SOON, OVERDUE
    created_at: datetime


@router.get("/payables/schedule")
async def get_payment_schedule(tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """
    Get payment schedule with due dates and urgency status.
    Shows all pending payments sorted by due date.
    """
    from datetime import timezone
    now = datetime.now(timezone.utc)
    
    result = await db.execute(
        select(PurchaseOrderModel)
        .options(selectinload(PurchaseOrderModel.supplier))
        .where(
            PurchaseOrderModel.tenant_id == tenant_id,
            PurchaseOrderModel.status.notin_(['PAID', 'CANCELLED', 'DRAFT'])
        )
        .order_by(PurchaseOrderModel.due_date.asc().nullsfirst())
    )
    orders = result.scalars().all()
    
    schedule = []
    for o in orders:
        # Handle timezone-aware datetime for created_at
        created = o.created_at or now
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        
        # Calculate due date based on payment terms if not set
        terms = o.payment_terms or 'NET30'
        
        if o.due_date:
            due = o.due_date
            if due.tzinfo is None:
                due = due.replace(tzinfo=timezone.utc)
        else:
            # Calculate based on terms
            term_days = {
                'IMMEDIATE': 0,
                'NET15': 15,
                'NET30': 30,
                'NET60': 60,
                'NET90': 90
            }
            days_offset = term_days.get(terms, 30)
            due = created + timedelta(days=days_offset)
        
        # Calculate days until due
        days_until = (due - now).days if due else 999
        
        # Determine status
        if days_until < 0:
            status = 'OVERDUE'
        elif days_until <= 7:
            status = 'DUE_SOON'
        else:
            status = 'ON_TIME'
        
        total = float(o.total_amount or 0)
        paid = float(o.paid_amount or 0)
        
        schedule.append(PaymentScheduleItem(
            po_id=o.id,
            po_code=o.code,
            supplier_name=o.supplier.name if o.supplier else "N/A",
            amount=total,
            paid_amount=paid,
            balance=total - paid,
            payment_terms=terms,
            due_date=due,
            days_until_due=days_until,
            status=status,
            created_at=o.created_at
        ))
    
    # Sort by urgency
    schedule.sort(key=lambda x: x.days_until_due)
    
    return {
        "schedule": schedule,
        "summary": {
            "total_count": len(schedule),
            "overdue_count": sum(1 for s in schedule if s.status == 'OVERDUE'),
            "due_soon_count": sum(1 for s in schedule if s.status == 'DUE_SOON'),
            "total_balance": sum(s.balance for s in schedule)
        }
    }


@router.put("/payables/{po_id}/payment-terms")
async def update_payment_terms(
    po_id: UUID,
    payment_terms: str,
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)
):
    """
    Update payment terms for a purchase order.
    Recalculates due date based on new terms.
    """
    now = datetime.now()
    
    result = await db.execute(
        select(PurchaseOrderModel).where(
            PurchaseOrderModel.id == po_id,
            PurchaseOrderModel.tenant_id == tenant_id
        )
    )
    order = result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    # Validate terms
    valid_terms = ['IMMEDIATE', 'NET15', 'NET30', 'NET60', 'NET90']
    if payment_terms not in valid_terms:
        raise HTTPException(status_code=400, detail=f"Invalid terms. Must be one of: {valid_terms}")
    
    # Update terms and recalculate due date
    order.payment_terms = payment_terms
    
    term_days = {
        'IMMEDIATE': 0, 'NET15': 15, 'NET30': 30, 'NET60': 60, 'NET90': 90
    }
    order.due_date = order.created_at + timedelta(days=term_days[payment_terms])
    
    await db.commit()
    
    return {
        "message": "Payment terms updated",
        "po_id": str(po_id),
        "payment_terms": payment_terms,
        "due_date": order.due_date.isoformat()
    }


# ============ REPORTS ============

class CashFlowItem(BaseModel):
    period: str
    receipts: float
    payments: float
    net_flow: float


class PnLItem(BaseModel):
    category: str
    amount: float


@router.get("/reports/cashflow")
async def get_cashflow_report(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    months: int = Query(6, le=12)
):
    """
    Cash Flow Statement - Thu/Chi theo tháng
    Shows receipts vs payments over time
    """
    now = datetime.now()
    result = []
    
    for i in range(months - 1, -1, -1):
        # Calculate month
        month_date = now - timedelta(days=i * 30)
        month_start = date(month_date.year, month_date.month, 1)
        if month_date.month == 12:
            month_end = date(month_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            month_end = date(month_date.year, month_date.month + 1, 1) - timedelta(days=1)
        
        # Receipts from Order Payments
        receipts_result = await db.execute(
            select(func.coalesce(func.sum(OrderPaymentModel.amount), 0))
            .where(
                func.date(OrderPaymentModel.payment_date) >= month_start,
                func.date(OrderPaymentModel.payment_date) <= month_end
            )
        )
        receipts = float(receipts_result.scalar() or 0)
        
        # Payments from Finance Transactions
        payments_result = await db.execute(
            select(func.coalesce(func.sum(FinanceTransactionModel.amount), 0))
            .where(
                FinanceTransactionModel.tenant_id == tenant_id,
                FinanceTransactionModel.type == 'PAYMENT',
                FinanceTransactionModel.transaction_date >= month_start,
                FinanceTransactionModel.transaction_date <= month_end
            )
        )
        payments = float(payments_result.scalar() or 0)
        
        result.append(CashFlowItem(
            period=month_start.strftime('%m/%Y'),
            receipts=receipts,
            payments=payments,
            net_flow=receipts - payments
        ))
    
    return result


@router.get("/reports/pnl")
async def get_pnl_report(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    year: int = Query(None),
    month: int = Query(None)
):
    """
    Profit & Loss Statement
    Revenue - COGS - Operating Expenses = Net Profit
    """
    now = datetime.now()
    if not year:
        year = now.year
    if not month:
        month = now.month
    
    month_start = date(year, month, 1)
    if month == 12:
        month_end = date(year + 1, 1, 1) - timedelta(days=1)
    else:
        month_end = date(year, month + 1, 1) - timedelta(days=1)
    
    # Revenue from Orders (paid amounts)
    revenue_result = await db.execute(
        select(func.coalesce(func.sum(OrderPaymentModel.amount), 0))
        .where(
            func.date(OrderPaymentModel.payment_date) >= month_start,
            func.date(OrderPaymentModel.payment_date) <= month_end
        )
    )
    revenue = float(revenue_result.scalar() or 0)
    
    # COGS from Procurement (materials)
    cogs_result = await db.execute(
        select(func.coalesce(func.sum(PurchaseOrderModel.total_amount), 0))
        .where(
            PurchaseOrderModel.tenant_id == tenant_id,
            PurchaseOrderModel.status.in_(['RECEIVED', 'PAID']),
            func.date(PurchaseOrderModel.created_at) >= month_start,
            func.date(PurchaseOrderModel.created_at) <= month_end
        )
    )
    cogs = float(cogs_result.scalar() or 0)
    
    # Operating expenses from Finance Transactions
    opex_result = await db.execute(
        select(func.coalesce(func.sum(FinanceTransactionModel.amount), 0))
        .where(
            FinanceTransactionModel.tenant_id == tenant_id,
            FinanceTransactionModel.type == 'PAYMENT',
            FinanceTransactionModel.category.in_(['SALARY', 'OPERATING', 'OTHER']),
            FinanceTransactionModel.transaction_date >= month_start,
            FinanceTransactionModel.transaction_date <= month_end
        )
    )
    opex = float(opex_result.scalar() or 0)
    
    gross_profit = revenue - cogs
    net_profit = gross_profit - opex
    
    return {
        "period": f"{month:02d}/{year}",
        "items": [
            {"category": "Doanh thu", "amount": revenue},
            {"category": "Giá vốn hàng bán (COGS)", "amount": -cogs},
            {"category": "Lợi nhuận gộp", "amount": gross_profit},
            {"category": "Chi phí vận hành", "amount": -opex},
            {"category": "Lợi nhuận ròng", "amount": net_profit}
        ],
        "summary": {
            "revenue": revenue,
            "cogs": cogs,
            "gross_profit": gross_profit,
            "opex": opex,
            "net_profit": net_profit,
            "gross_margin": round(gross_profit / revenue * 100, 2) if revenue > 0 else 0,
            "net_margin": round(net_profit / revenue * 100, 2) if revenue > 0 else 0
        }
    }


# ============ PHASE 5A: EVENT-BASED P&L ============

class OrderPnLResponse(BaseModel):
    order_id: UUID
    order_code: str
    customer_name: str
    event_date: Optional[datetime]
    
    # Revenue
    final_amount: float
    paid_amount: float
    
    # Costs
    estimated_cogs: float  # From quote items
    actual_cogs: float     # From procurement
    labor_cost: float      # From HR timesheets (if available)
    overhead: float        # Allocated overhead
    total_costs: float
    
    # Profits
    gross_profit: float
    net_profit: float
    gross_margin: float
    net_margin: float
    
    # Variance
    estimated_profit: float
    profit_variance: float


class ProfitabilitySummary(BaseModel):
    total_orders: int
    total_revenue: float
    total_costs: float
    total_profit: float
    avg_margin: float
    orders: List[OrderPnLResponse]


@router.get("/orders/{order_id}/pnl", response_model=OrderPnLResponse)
async def get_order_pnl(order_id: UUID, tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)):
    """
    Get P&L breakdown for a specific order.
    Shows revenue, costs, and profit margin per event.
    """
    # Get order details
    order_result = await db.execute(
        select(OrderModel).where(
            OrderModel.id == order_id,
            OrderModel.tenant_id == tenant_id
        )
    )
    order = order_result.scalar_one_or_none()
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Revenue
    final_amount = float(order.final_amount or 0)
    paid_amount = float(order.paid_amount or 0)
    
    # Estimated COGS from quote items (food cost estimate ~40%)
    estimated_cogs = final_amount * 0.40
    
    # Actual COGS from linked purchase orders (if any)
    # For now, estimate based on standard food cost ratio
    actual_cogs = final_amount * 0.38  # Slightly better than estimate
    
    # Labor cost (from HR timesheets with actual hourly rates - SOL-2 Fix)
    labor_cost = 0.0
    try:
        from backend.modules.hr.domain.models import TimesheetModel, EmployeeModel
        
        # Get timesheets linked to this order with employee hourly rates
        timesheet_result = await db.execute(
            select(
                TimesheetModel.total_hours,
                EmployeeModel.hourly_rate
            ).select_from(TimesheetModel)
            .outerjoin(EmployeeModel, TimesheetModel.employee_id == EmployeeModel.id)
            .where(
                TimesheetModel.order_id == order_id,
                TimesheetModel.tenant_id == tenant_id
            )
        )
        rows = timesheet_result.all()
        
        if rows:
            for hours, rate in rows:
                h = float(hours or 0)
                r = float(rate or 50000)  # Fallback to 50k if no rate set
                labor_cost += h * r
        else:
            # Fallback: Try older StaffAssignment-based lookup
            from backend.modules.hr.domain.models import StaffAssignmentModel
            assignment_result = await db.execute(
                select(func.coalesce(func.sum(
                    func.coalesce(TimesheetModel.total_hours, 8) * 
                    func.coalesce(EmployeeModel.hourly_rate, 50000)
                ), 0))
                .select_from(StaffAssignmentModel)
                .outerjoin(TimesheetModel, TimesheetModel.assignment_id == StaffAssignmentModel.id)
                .outerjoin(EmployeeModel, StaffAssignmentModel.employee_id == EmployeeModel.id)
                .where(
                    StaffAssignmentModel.event_id == order_id,
                    StaffAssignmentModel.tenant_id == tenant_id
                )
            )
            labor_cost = float(assignment_result.scalar() or 0)
            
            # If still no data, estimate based on revenue
            if labor_cost == 0:
                labor_cost = final_amount * 0.15  # Last resort: 15% estimate
    except Exception as e:
        # HR module not found or error
        labor_cost = final_amount * 0.15  # Estimate 15% of revenue
    
    # Overhead allocation (rent, utilities, admin - ~10%)
    overhead = final_amount * 0.10
    
    # Calculate totals
    total_costs = actual_cogs + labor_cost + overhead
    gross_profit = final_amount - actual_cogs
    net_profit = final_amount - total_costs
    
    # Margins
    gross_margin = round(gross_profit / final_amount * 100, 2) if final_amount > 0 else 0
    net_margin = round(net_profit / final_amount * 100, 2) if final_amount > 0 else 0
    
    # Variance from estimate
    estimated_profit = final_amount - estimated_cogs - (final_amount * 0.25)  # Est 25% other costs
    profit_variance = net_profit - estimated_profit
    
    return OrderPnLResponse(
        order_id=order.id,
        order_code=order.code or str(order.id)[:8],
        customer_name=order.customer_name or "Unknown",
        event_date=order.event_date,
        final_amount=final_amount,
        paid_amount=paid_amount,
        estimated_cogs=estimated_cogs,
        actual_cogs=actual_cogs,
        labor_cost=labor_cost,
        overhead=overhead,
        total_costs=total_costs,
        gross_profit=gross_profit,
        net_profit=net_profit,
        gross_margin=gross_margin,
        net_margin=net_margin,
        estimated_profit=estimated_profit,
        profit_variance=profit_variance
    )


@router.get("/orders/profitability", response_model=ProfitabilitySummary)
async def get_orders_profitability(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    limit: int = Query(20, le=100)
):
    """
    Get profitability analysis for multiple orders.
    Returns ranking by profit margin.
    """
    query = select(OrderModel).where(
        OrderModel.tenant_id == tenant_id,
        OrderModel.status.in_(['COMPLETED', 'PAID', 'CONFIRMED'])
    )
    
    if start_date:
        query = query.where(func.date(OrderModel.event_date) >= date.fromisoformat(start_date))
    if end_date:
        query = query.where(func.date(OrderModel.event_date) <= date.fromisoformat(end_date))
    
    query = query.order_by(desc(OrderModel.event_date)).limit(limit)
    
    result = await db.execute(query)
    orders = result.scalars().all()
    
    # Pre-fetch labor costs for all orders (SOL-2: Batch query for performance)
    order_ids = [o.id for o in orders]
    labor_costs_map = {}
    
    if order_ids:
        try:
            from backend.modules.hr.domain.models import TimesheetModel, EmployeeModel
            
            labor_result = await db.execute(
                select(
                    TimesheetModel.order_id,
                    func.sum(
                        func.coalesce(TimesheetModel.total_hours, 0) * 
                        func.coalesce(EmployeeModel.hourly_rate, 50000)
                    ).label('labor_cost')
                ).select_from(TimesheetModel)
                .outerjoin(EmployeeModel, TimesheetModel.employee_id == EmployeeModel.id)
                .where(
                    TimesheetModel.order_id.in_(order_ids),
                    TimesheetModel.tenant_id == tenant_id
                )
                .group_by(TimesheetModel.order_id)
            )
            for order_id, cost in labor_result.all():
                labor_costs_map[order_id] = float(cost or 0)
        except Exception:
            pass  # Will use fallback estimate
    
    order_pnls = []
    total_revenue = 0.0
    total_costs = 0.0
    total_profit = 0.0
    
    for order in orders:
        final_amount = float(order.final_amount or 0)
        paid_amount = float(order.paid_amount or 0)
        
        # Cost calculations
        estimated_cogs = final_amount * 0.40
        actual_cogs = final_amount * 0.38
        
        # Labor cost from timesheets or fallback to estimate
        labor_cost = labor_costs_map.get(order.id, 0)
        if labor_cost == 0:
            labor_cost = final_amount * 0.15  # Fallback estimate
        
        overhead = final_amount * 0.10
        costs = actual_cogs + labor_cost + overhead
        
        gross_profit = final_amount - actual_cogs
        net_profit = final_amount - costs
        
        gross_margin = round(gross_profit / final_amount * 100, 2) if final_amount > 0 else 0
        net_margin = round(net_profit / final_amount * 100, 2) if final_amount > 0 else 0
        
        estimated_profit = final_amount * 0.35  # Standard 35% margin target
        
        order_pnls.append(OrderPnLResponse(
            order_id=order.id,
            order_code=order.code or str(order.id)[:8],
            customer_name=order.customer_name or "Unknown",
            event_date=order.event_date,
            final_amount=final_amount,
            paid_amount=paid_amount,
            estimated_cogs=estimated_cogs,
            actual_cogs=actual_cogs,
            labor_cost=labor_cost,
            overhead=overhead,
            total_costs=costs,
            gross_profit=gross_profit,
            net_profit=net_profit,
            gross_margin=gross_margin,
            net_margin=net_margin,
            estimated_profit=estimated_profit,
            profit_variance=net_profit - estimated_profit
        ))
        
        total_revenue += final_amount
        total_costs += costs
        total_profit += net_profit
    
    avg_margin = round(total_profit / total_revenue * 100, 2) if total_revenue > 0 else 0
    
    return ProfitabilitySummary(
        total_orders=len(order_pnls),
        total_revenue=total_revenue,
        total_costs=total_costs,
        total_profit=total_profit,
        avg_margin=avg_margin,
        orders=order_pnls
    )


# ============ PHASE 5C: HR → FINANCE INTEGRATION ============

class LaborEntryRequest(BaseModel):
    start_date: str  # YYYY-MM-DD
    end_date: str    # YYYY-MM-DD
    hourly_rate: float = 50000  # Default 50k VND/hour


class LaborEntryResult(BaseModel):
    entries_created: int
    total_amount: float
    transaction_ids: List[str]


@router.post("/auto-entries/from-timesheets", response_model=LaborEntryResult)
async def create_labor_entries_from_timesheets(
    data: LaborEntryRequest,
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)
):
    """
    Auto-generate finance transactions from HR timesheets.
    Creates PAYMENT transactions for labor costs.
    """
    from backend.modules.hr.domain.models import TimesheetModel, EmployeeModel
    
    start = date.fromisoformat(data.start_date)
    end = date.fromisoformat(data.end_date)
    
    # Get approved timesheets in date range
    result = await db.execute(
        select(
            TimesheetModel,
            EmployeeModel.full_name
        ).outerjoin(
            EmployeeModel, TimesheetModel.employee_id == EmployeeModel.id
        ).where(
            TimesheetModel.tenant_id == tenant_id,
            TimesheetModel.work_date >= start,
            TimesheetModel.work_date <= end,
            TimesheetModel.status == 'APPROVED'
        )
    )
    rows = result.all()
    
    if not rows:
        return LaborEntryResult(entries_created=0, total_amount=0, transaction_ids=[])
    
    # Generate transaction code
    now = datetime.now()
    code_prefix = f"CHI-LAB-{now.strftime('%Y%m%d')}"
    
    transaction_ids = []
    total_amount = 0.0
    
    for ts, emp_name in rows:
        hours = float(ts.total_hours or 0)
        amount = hours * data.hourly_rate
        
        if amount > 0:
            # Create finance transaction
            txn = FinanceTransactionModel(
                tenant_id=tenant_id,
                code=f"{code_prefix}-{len(transaction_ids)+1:03d}",
                type='PAYMENT',
                category='SALARY',
                amount=Decimal(str(amount)),
                payment_method='BANK_TRANSFER',
                description=f"Lương nhân viên: {emp_name} - {ts.work_date}",
                transaction_date=ts.work_date,
                reference_id=ts.employee_id,
                reference_type='TIMESHEET'
            )
            db.add(txn)
            transaction_ids.append(str(txn.id) if txn.id else "pending")
            total_amount += amount
    
    await db.commit()
    
    return LaborEntryResult(
        entries_created=len(transaction_ids),
        total_amount=total_amount,
        transaction_ids=transaction_ids
    )


@router.get("/labor-costs/by-event")
async def get_labor_costs_by_event(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    event_id: Optional[UUID] = Query(None)
):
    """
    Get labor costs breakdown by event/order.
    Aggregates timesheet hours for staff assigned to events.
    """
    from backend.modules.hr.domain.models import StaffAssignmentModel, TimesheetModel, EmployeeModel
    
    query = select(
        StaffAssignmentModel.event_id,
        EmployeeModel.full_name,
        EmployeeModel.hourly_rate,
        func.sum(TimesheetModel.total_hours).label('total_hours')
    ).select_from(
        StaffAssignmentModel
    ).outerjoin(
        TimesheetModel, TimesheetModel.assignment_id == StaffAssignmentModel.id
    ).outerjoin(
        EmployeeModel, StaffAssignmentModel.employee_id == EmployeeModel.id
    ).where(
        StaffAssignmentModel.tenant_id == tenant_id
    ).group_by(
        StaffAssignmentModel.event_id,
        EmployeeModel.full_name,
        EmployeeModel.hourly_rate
    )
    
    if event_id:
        query = query.where(StaffAssignmentModel.event_id == event_id)
    
    result = await db.execute(query)
    rows = result.all()
    
    labor_data = []
    for eid, name, rate, hours in rows:
        h = float(hours or 0)
        r = float(rate or 50000)
        labor_data.append({
            "event_id": str(eid) if eid else None,
            "employee_name": name,
            "hours": h,
            "hourly_rate": r,
            "cost": h * r
        })
    
    total_cost = sum(item['cost'] for item in labor_data)
    
    return {
        "labor_entries": labor_data,
        "total_hours": sum(item['hours'] for item in labor_data),
        "total_cost": total_cost
    }


# ============ PHASE 5D: CASH FLOW FORECAST ============

class ForecastItem(BaseModel):
    period: str
    expected_receipts: float
    expected_payments: float
    net_forecast: float
    cumulative: float


@router.get("/forecast/cashflow")
async def get_cashflow_forecast(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    weeks: int = Query(4, le=12)
):
    """
    Cash flow forecast based on scheduled events and pending payables.
    Predicts receipts from upcoming orders and payments to suppliers.
    """
    now = datetime.now()
    forecasts = []
    cumulative = 0.0
    
    for w in range(weeks):
        week_start = now + timedelta(weeks=w)
        week_end = week_start + timedelta(days=7)
        
        # Expected receipts from orders with event dates in this week
        receipts_result = await db.execute(
            select(func.coalesce(func.sum(OrderModel.balance_amount), 0))
            .where(
                OrderModel.tenant_id == tenant_id,
                OrderModel.event_date >= week_start,
                OrderModel.event_date < week_end,
                OrderModel.status.in_(['CONFIRMED', 'IN_PROGRESS'])
            )
        )
        expected_receipts = float(receipts_result.scalar() or 0)
        
        # Expected payments from purchase orders due this week
        payments_result = await db.execute(
            select(func.coalesce(func.sum(PurchaseOrderModel.total_amount - func.coalesce(PurchaseOrderModel.paid_amount, 0)), 0))
            .where(
                PurchaseOrderModel.tenant_id == tenant_id,
                PurchaseOrderModel.due_date >= week_start,
                PurchaseOrderModel.due_date < week_end,
                PurchaseOrderModel.status.notin_(['PAID', 'CANCELLED'])
            )
        )
        expected_payments = float(payments_result.scalar() or 0)
        
        net = expected_receipts - expected_payments
        cumulative += net
        
        forecasts.append(ForecastItem(
            period=f"Tuần {w+1} ({week_start.strftime('%d/%m')} - {week_end.strftime('%d/%m')})",
            expected_receipts=expected_receipts,
            expected_payments=expected_payments,
            net_forecast=net,
            cumulative=cumulative
        ))
    
    # Summary
    total_receipts = sum(f.expected_receipts for f in forecasts)
    total_payments = sum(f.expected_payments for f in forecasts)
    
    return {
        "forecast": forecasts,
        "summary": {
            "total_expected_receipts": total_receipts,
            "total_expected_payments": total_payments,
            "net_position": total_receipts - total_payments,
            "forecast_weeks": weeks
        }
    }


@router.get("/forecast/upcoming-receipts")
async def get_upcoming_receipts(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    days: int = Query(30, le=90)
):
    """
    List of expected receipts from confirmed orders.
    Sorted by event date.
    """
    now = datetime.now()
    future = now + timedelta(days=days)
    
    result = await db.execute(
        select(OrderModel).where(
            OrderModel.tenant_id == tenant_id,
            OrderModel.event_date >= now,
            OrderModel.event_date <= future,
            OrderModel.balance_amount > 0,
            OrderModel.status.in_(['CONFIRMED', 'IN_PROGRESS'])
        ).order_by(OrderModel.event_date.asc())
    )
    orders = result.scalars().all()
    
    upcoming = []
    for o in orders:
        upcoming.append({
            "order_id": str(o.id),
            "order_code": o.code,
            "customer_name": o.customer_name,
            "event_date": o.event_date.isoformat() if o.event_date else None,
            "final_amount": float(o.final_amount or 0),
            "paid_amount": float(o.paid_amount or 0),
            "balance": float(o.balance_amount or 0),
            "days_until_event": (o.event_date - now).days if o.event_date else 0
        })
    
    return {
        "receipts": upcoming,
        "total_expected": sum(r['balance'] for r in upcoming),
        "count": len(upcoming)
    }


# ============ HR-FINANCE INTEGRATION ============

from backend.modules.hr.domain.models import PayrollPeriodModel

@router.post("/auto-entries/from-payroll/{period_id}")
async def create_salary_expense_from_payroll(
    period_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db)
):
    """
    Create finance expense transaction from approved payroll period.
    Automatically generates PAYMENT transaction for salary category.
    """
    import random
    
    # Get payroll period
    period_result = await db.execute(
        select(PayrollPeriodModel).where(
            PayrollPeriodModel.id == period_id,
            PayrollPeriodModel.tenant_id == tenant_id
        )
    )
    period = period_result.scalar_one_or_none()
    
    if not period:
        raise HTTPException(status_code=404, detail="Payroll period not found")
    
    if period.status not in ['APPROVED', 'PAID']:
        raise HTTPException(status_code=400, detail="Payroll must be approved before creating expense")
    
    # Check if expense already exists
    existing = await db.execute(
        select(FinanceTransactionModel).where(
            FinanceTransactionModel.tenant_id == tenant_id,
            FinanceTransactionModel.reference_id == period_id,
            FinanceTransactionModel.reference_type == 'PAYROLL'
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Salary expense already created for this period")
    
    # Create expense transaction
    now = datetime.now()
    random_suffix = random.randint(1000, 9999)
    code = f"CHI-{now.strftime('%Y%m')}-{random_suffix}"
    
    salary_transaction = FinanceTransactionModel(
        tenant_id=tenant_id,
        code=code,
        type='PAYMENT',
        category='SALARY',
        amount=Decimal(str(period.total_net or 0)),
        payment_method='BANK_TRANSFER',
        description=f"Chi lương kỳ {period.period_name} ({int(period.total_employees or 0)} nhân viên)",
        transaction_date=date.today(),
        reference_id=period_id,
        reference_type='PAYROLL'
    )
    
    db.add(salary_transaction)
    
    # Update payroll period status to PAID
    period.status = 'PAID'
    period.paid_at = now
    
    await db.commit()
    
    return {
        "message": f"Đã tạo chi phí lương {period.period_name}",
        "transaction_code": code,
        "amount": float(period.total_net or 0),
        "employees": int(period.total_employees or 0)
    }


@router.get("/labor-costs/summary")
async def get_labor_cost_summary(
    tenant_id: UUID = Depends(get_current_tenant), db: AsyncSession = Depends(get_db),
    months: int = Query(6, le=12)
):
    """
    Get labor cost summary from payroll data.
    Shows salary expenses by month.
    """
    # Get salary transactions
    result = await db.execute(
        select(FinanceTransactionModel).where(
            FinanceTransactionModel.tenant_id == tenant_id,
            FinanceTransactionModel.category == 'SALARY',
            FinanceTransactionModel.type == 'PAYMENT'
        ).order_by(FinanceTransactionModel.transaction_date.desc())
        .limit(months)
    )
    transactions = result.scalars().all()
    
    # Group by month
    by_month = {}
    for t in transactions:
        key = t.transaction_date.strftime('%m/%Y')
        if key not in by_month:
            by_month[key] = 0
        by_month[key] += float(t.amount)
    
    return {
        "labor_costs": [
            {"period": k, "amount": v} for k, v in by_month.items()
        ],
        "total": sum(by_month.values()),
        "average": sum(by_month.values()) / len(by_month) if by_month else 0
    }


# ============ REPORT EXPORT ============

from fastapi.responses import StreamingResponse
import io

@router.get("/reports/export")
async def export_finance_report(
    format: str = Query("excel", description="Export format: excel or pdf"),
    from_date: str = Query(..., description="Start date (YYYY-MM-DD)"),
    to_date: str = Query(..., description="End date (YYYY-MM-DD)"),
    report_type: str = Query("transactions", description="Report type: transactions, receivables, summary"),
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Export finance report in Excel or PDF format.
    """
    from datetime import datetime as dt
    
    try:
        start_date = dt.strptime(from_date, "%Y-%m-%d").date()
        end_date = dt.strptime(to_date, "%Y-%m-%d").date()
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD")
    
    if format.lower() == "excel":
        return await _export_excel(db, tenant_id, start_date, end_date, report_type)
    elif format.lower() == "pdf":
        return await _export_pdf(db, tenant_id, start_date, end_date, report_type)
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'excel' or 'pdf'")


async def _export_excel(db: AsyncSession, tenant_id: UUID, start_date: date, end_date: date, report_type: str):
    """Generate Excel report using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        raise HTTPException(
            status_code=500, 
            detail="openpyxl not installed. Run: pip install openpyxl"
        )
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo Cáo Tài Chính"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"BÁO CÁO TÀI CHÍNH - {report_type.upper()}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    ws['A2'] = f"Từ ngày: {start_date.strftime('%d/%m/%Y')} đến ngày: {end_date.strftime('%d/%m/%Y')}"
    ws.merge_cells('A2:F2')
    
    if report_type == "transactions":
        # Query transactions
        result = await db.execute(
            select(FinanceTransactionModel)
            .where(
                FinanceTransactionModel.tenant_id == tenant_id,
                FinanceTransactionModel.transaction_date >= start_date,
                FinanceTransactionModel.transaction_date <= end_date
            )
            .order_by(FinanceTransactionModel.transaction_date.desc())
        )
        transactions = result.scalars().all()
        
        # Headers
        headers = ["Mã GD", "Ngày", "Loại", "Danh mục", "Mô tả", "Số tiền"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Data
        total_receipt = 0
        total_payment = 0
        for row, t in enumerate(transactions, 5):
            ws.cell(row=row, column=1, value=t.code).border = thin_border
            ws.cell(row=row, column=2, value=t.transaction_date.strftime('%d/%m/%Y')).border = thin_border
            ws.cell(row=row, column=3, value="Thu" if t.type == "RECEIPT" else "Chi").border = thin_border
            ws.cell(row=row, column=4, value=t.category or "").border = thin_border
            ws.cell(row=row, column=5, value=t.description or "").border = thin_border
            ws.cell(row=row, column=6, value=float(t.amount)).border = thin_border
            ws.cell(row=row, column=6).number_format = '#,##0'
            
            if t.type == "RECEIPT":
                total_receipt += float(t.amount)
            else:
                total_payment += float(t.amount)
        
        # Summary
        summary_row = len(transactions) + 6
        ws.cell(row=summary_row, column=5, value="Tổng Thu:").font = Font(bold=True)
        ws.cell(row=summary_row, column=6, value=total_receipt).number_format = '#,##0'
        ws.cell(row=summary_row + 1, column=5, value="Tổng Chi:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=6, value=total_payment).number_format = '#,##0'
        ws.cell(row=summary_row + 2, column=5, value="Chênh lệch:").font = Font(bold=True)
        ws.cell(row=summary_row + 2, column=6, value=total_receipt - total_payment).number_format = '#,##0'
    
    elif report_type == "receivables":
        # Query unpaid orders
        result = await db.execute(
            select(OrderModel)
            .where(
                OrderModel.tenant_id == tenant_id,
                OrderModel.balance_amount > 0,
                OrderModel.status.in_(['CONFIRMED', 'COMPLETED'])
            )
            .order_by(OrderModel.event_date.desc())
        )
        orders = result.scalars().all()
        
        headers = ["Mã ĐH", "Khách hàng", "SĐT", "Ngày sự kiện", "Tổng tiền", "Đã thu", "Còn nợ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        total_receivable = 0
        for row, o in enumerate(orders, 5):
            ws.cell(row=row, column=1, value=o.code)
            ws.cell(row=row, column=2, value=o.customer_name)
            ws.cell(row=row, column=3, value=o.customer_phone)
            ws.cell(row=row, column=4, value=o.event_date.strftime('%d/%m/%Y') if o.event_date else "")
            ws.cell(row=row, column=5, value=float(o.final_amount)).number_format = '#,##0'
            ws.cell(row=row, column=6, value=float(o.paid_amount)).number_format = '#,##0'
            ws.cell(row=row, column=7, value=float(o.balance_amount)).number_format = '#,##0'
            total_receivable += float(o.balance_amount)
        
        summary_row = len(orders) + 6
        ws.cell(row=summary_row, column=6, value="Tổng công nợ:").font = Font(bold=True)
        ws.cell(row=summary_row, column=7, value=total_receivable).number_format = '#,##0'
    
    # Auto-width columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"bao_cao_tai_chinh_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def _export_pdf(db: AsyncSession, tenant_id: UUID, start_date: date, end_date: date, report_type: str):
    """Generate PDF report - simplified HTML-based approach"""
    # For now, return a simple HTML that can be printed as PDF
    # In production, use reportlab or weasyprint
    
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Báo Cáo Tài Chính</title>
        <style>
            body {{ font-family: Arial, sans-serif; margin: 40px; }}
            h1 {{ text-align: center; color: #333; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 20px; }}
            th {{ background-color: #4472C4; color: white; padding: 10px; text-align: left; }}
            td {{ padding: 8px; border-bottom: 1px solid #ddd; }}
            .summary {{ margin-top: 20px; font-weight: bold; }}
            .amount {{ text-align: right; }}
            @media print {{ body {{ margin: 0; }} }}
        </style>
    </head>
    <body>
        <h1>BÁO CÁO TÀI CHÍNH - {report_type.upper()}</h1>
        <p>Từ ngày: {start_date.strftime('%d/%m/%Y')} đến ngày: {end_date.strftime('%d/%m/%Y')}</p>
    """
    
    if report_type == "transactions":
        result = await db.execute(
            select(FinanceTransactionModel)
            .where(
                FinanceTransactionModel.tenant_id == tenant_id,
                FinanceTransactionModel.transaction_date >= start_date,
                FinanceTransactionModel.transaction_date <= end_date
            )
            .order_by(FinanceTransactionModel.transaction_date.desc())
        )
        transactions = result.scalars().all()
        
        html_content += """
        <table>
            <tr><th>Mã GD</th><th>Ngày</th><th>Loại</th><th>Danh mục</th><th>Mô tả</th><th>Số tiền</th></tr>
        """
        
        total_receipt = 0
        total_payment = 0
        for t in transactions:
            type_label = "Thu" if t.type == "RECEIPT" else "Chi"
            amount = float(t.amount)
            if t.type == "RECEIPT":
                total_receipt += amount
            else:
                total_payment += amount
            
            html_content += f"""
            <tr>
                <td>{t.code}</td>
                <td>{t.transaction_date.strftime('%d/%m/%Y')}</td>
                <td>{type_label}</td>
                <td>{t.category or ''}</td>
                <td>{t.description or ''}</td>
                <td class="amount">{amount:,.0f}</td>
            </tr>
            """
        
        html_content += f"""
        </table>
        <div class="summary">
            <p>Tổng Thu: {total_receipt:,.0f} VND</p>
            <p>Tổng Chi: {total_payment:,.0f} VND</p>
            <p>Chênh lệch: {total_receipt - total_payment:,.0f} VND</p>
        </div>
        """
    
    html_content += """
    </body>
    </html>
    """
    
    buffer = io.BytesIO(html_content.encode('utf-8'))
    buffer.seek(0)
    
    filename = f"bao_cao_tai_chinh_{start_date.strftime('%Y%m%d')}_{end_date.strftime('%Y%m%d')}.html"
    
    return StreamingResponse(
        buffer,
        media_type="text/html",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ============ PHASE 3: PERIOD CLOSING (FIN-004) ============

from backend.modules.finance.domain.models import AccountingPeriodModel


class PeriodResponse(BaseModel):
    id: UUID
    name: str
    period_type: str
    start_date: date
    end_date: date
    status: str
    closed_at: Optional[datetime] = None
    closing_total_debit: Optional[float] = None
    closing_total_credit: Optional[float] = None
    closing_retained_earnings: Optional[float] = None
    notes: Optional[str] = None
    created_at: datetime

    class Config:
        from_attributes = True


class CreatePeriodRequest(BaseModel):
    name: str
    period_type: str = 'MONTHLY'  # MONTHLY, QUARTERLY, YEARLY
    start_date: date
    end_date: date
    notes: Optional[str] = None


@router.get("/periods", response_model=List[PeriodResponse])
async def list_accounting_periods(
    status: Optional[str] = None,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """List all accounting periods, optionally filtered by status."""
    query = select(AccountingPeriodModel).where(
        AccountingPeriodModel.tenant_id == tenant_id
    )
    
    if status:
        query = query.where(AccountingPeriodModel.status == status)
    
    query = query.order_by(AccountingPeriodModel.start_date.desc())
    
    result = await db.execute(query)
    periods = result.scalars().all()
    
    return periods


@router.post("/periods", response_model=PeriodResponse)
async def create_accounting_period(
    data: CreatePeriodRequest,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Create a new accounting period."""
    # Validate period type
    valid_types = ['MONTHLY', 'QUARTERLY', 'YEARLY']
    if data.period_type not in valid_types:
        raise HTTPException(status_code=400, detail=f"Invalid period type. Must be one of: {valid_types}")
    
    # Validate dates
    if data.end_date < data.start_date:
        raise HTTPException(status_code=400, detail="End date must be after start date")
    
    # Check for overlapping periods
    overlap_result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.tenant_id == tenant_id,
            AccountingPeriodModel.start_date <= data.end_date,
            AccountingPeriodModel.end_date >= data.start_date
        )
    )
    if overlap_result.scalar_one_or_none():
        raise HTTPException(status_code=400, detail="Period overlaps with existing period")
    
    # Create period
    period = AccountingPeriodModel(
        tenant_id=tenant_id,
        name=data.name,
        period_type=data.period_type,
        start_date=data.start_date,
        end_date=data.end_date,
        notes=data.notes,
        status='OPEN'
    )
    
    db.add(period)
    await db.commit()
    await db.refresh(period)
    
    return period


@router.get("/periods/current")
async def get_current_period(
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """Get the current open period."""
    today = date.today()
    
    # First, try to find an OPEN period for today
    result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.tenant_id == tenant_id,
            AccountingPeriodModel.status == 'OPEN',
            AccountingPeriodModel.start_date <= today,
            AccountingPeriodModel.end_date >= today
        ).order_by(AccountingPeriodModel.created_at.desc())
    )
    period = result.scalar_one_or_none()
    
    if period:
        return period
    
    # No open period found - check if ANY period exists for this date range
    from calendar import monthrange
    first_day = today.replace(day=1)
    last_day = today.replace(day=monthrange(today.year, today.month)[1])
    
    existing_result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.tenant_id == tenant_id,
            AccountingPeriodModel.start_date == first_day,
            AccountingPeriodModel.end_date == last_day
        ).order_by(AccountingPeriodModel.created_at.desc())
    )
    existing_period = existing_result.scalar_one_or_none()
    
    if existing_period:
        # Return existing period (might be CLOSED - caller can decide to reopen)
        return existing_period
    
    # No period exists at all - auto-create
    period = AccountingPeriodModel(
        tenant_id=tenant_id,
        name=f"Tháng {today.month:02d}/{today.year}",
        period_type='MONTHLY',
        start_date=first_day,
        end_date=last_day,
        status='OPEN'
    )
    db.add(period)
    await db.commit()
    await db.refresh(period)
    
    return period


@router.post("/periods/{period_id}/close")
async def close_accounting_period(
    period_id: UUID,
    notes: Optional[str] = None,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Close an accounting period.
    - Validates all journals in period are POSTED
    - Calculates closing balances
    - Sets status to CLOSED
    """
    # Get period
    result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.id == period_id,
            AccountingPeriodModel.tenant_id == tenant_id
        )
    )
    period = result.scalar_one_or_none()
    
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    
    if period.status == 'CLOSED':
        raise HTTPException(status_code=400, detail="Period is already closed")
    
    # Check for DRAFT journals in period
    draft_result = await db.execute(
        select(func.count(JournalModel.id)).where(
            JournalModel.tenant_id == tenant_id,
            JournalModel.status == 'DRAFT',
            func.date(JournalModel.date) >= period.start_date,
            func.date(JournalModel.date) <= period.end_date
        )
    )
    draft_count = draft_result.scalar() or 0
    
    if draft_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot close period. {draft_count} draft journal(s) exist. Post or delete them first."
        )
    
    # Calculate closing balances
    balance_result = await db.execute(
        select(
            func.coalesce(func.sum(JournalLineModel.debit), 0).label('total_debit'),
            func.coalesce(func.sum(JournalLineModel.credit), 0).label('total_credit')
        )
        .join(JournalModel, JournalLineModel.journal_id == JournalModel.id)
        .where(
            JournalLineModel.tenant_id == tenant_id,
            JournalModel.status == 'POSTED',
            func.date(JournalModel.date) >= period.start_date,
            func.date(JournalModel.date) <= period.end_date
        )
    )
    balances = balance_result.one()
    
    # Update period
    period.status = 'CLOSED'
    period.closed_at = datetime.now()
    period.closing_total_debit = float(balances.total_debit)
    period.closing_total_credit = float(balances.total_credit)
    period.closing_retained_earnings = float(balances.total_credit - balances.total_debit)  # Simplified
    
    if notes:
        period.notes = (period.notes or '') + f"\n[Đóng kỳ]: {notes}"
    
    await db.commit()
    await db.refresh(period)
    
    return {
        "message": f"Đã đóng kỳ kế toán '{period.name}' thành công",
        "period": PeriodResponse.model_validate(period),
        "closing_balances": {
            "total_debit": period.closing_total_debit,
            "total_credit": period.closing_total_credit,
            "retained_earnings": period.closing_retained_earnings
        }
    }


class ReopenPeriodRequest(BaseModel):
    """Request body for reopening a period"""
    reason: str = "Mở lại để chỉnh sửa"

@router.post("/periods/{period_id}/reopen")
async def reopen_accounting_period(
    period_id: UUID,
    request: ReopenPeriodRequest = ReopenPeriodRequest(),
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Reopen a closed accounting period.
    Requires a reason for audit trail.
    """
    # Get period
    result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.id == period_id,
            AccountingPeriodModel.tenant_id == tenant_id
        )
    )
    period = result.scalar_one_or_none()
    
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    
    if period.status != 'CLOSED':
        raise HTTPException(status_code=400, detail="Only closed periods can be reopened")
    
    # Reopen
    period.status = 'OPEN'
    period.notes = (period.notes or '') + f"\n[Mở lại kỳ {datetime.now().strftime('%d/%m/%Y %H:%M')}]: {request.reason}"
    period.closed_at = None
    
    await db.commit()
    await db.refresh(period)
    
    return {
        "message": f"Đã mở lại kỳ kế toán '{period.name}'",
        "period": PeriodResponse.model_validate(period)
    }


@router.delete("/periods/{period_id}")
async def delete_accounting_period(
    period_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Delete an accounting period.
    Only OPEN periods with no journals can be deleted.
    """
    # Get period
    result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.id == period_id,
            AccountingPeriodModel.tenant_id == tenant_id
        )
    )
    period = result.scalar_one_or_none()
    
    if not period:
        raise HTTPException(status_code=404, detail="Period not found")
    
    if period.status == 'CLOSED':
        raise HTTPException(status_code=400, detail="Cannot delete closed period")
    
    # Check for journals in period
    journal_result = await db.execute(
        select(func.count(JournalModel.id)).where(
            JournalModel.tenant_id == tenant_id,
            func.date(JournalModel.date) >= period.start_date,
            func.date(JournalModel.date) <= period.end_date
        )
    )
    journal_count = journal_result.scalar() or 0
    
    if journal_count > 0:
        raise HTTPException(
            status_code=400,
            detail=f"Cannot delete period with {journal_count} journal(s). Delete journals first."
        )
    
    period_name = period.name
    await db.delete(period)
    await db.commit()
    
    return {"message": f"Đã xóa kỳ kế toán '{period_name}'"}


@router.post("/periods/cleanup-duplicates")
async def cleanup_duplicate_periods(
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Cleanup duplicate accounting periods.
    Keeps the first period (by created_at desc) for each date range, deletes the rest.
    Only deletes OPEN duplicate periods (not CLOSED ones).
    """
    # Get all periods grouped by date range
    all_periods_result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.tenant_id == tenant_id
        ).order_by(
            AccountingPeriodModel.start_date,
            AccountingPeriodModel.end_date,
            AccountingPeriodModel.created_at.desc()
        )
    )
    all_periods = all_periods_result.scalars().all()
    
    # Group by (start_date, end_date)
    period_groups = {}
    for p in all_periods:
        key = (p.start_date, p.end_date)
        if key not in period_groups:
            period_groups[key] = []
        period_groups[key].append(p)
    
    deleted_count = 0
    kept_periods = []
    deleted_ids = []
    
    for key, periods in period_groups.items():
        if len(periods) > 1:
            # Keep the first one (most recent created_at due to ordering)
            keep = periods[0]
            kept_periods.append(keep.name)
            
            # Delete the rest if they are OPEN (not CLOSED)
            # Note: We don't check for journals because journals are tied to 
            # date ranges, not specific period IDs
            for dup in periods[1:]:
                if dup.status == 'OPEN':
                    deleted_ids.append(str(dup.id))
                    await db.delete(dup)
                    deleted_count += 1
    
    await db.commit()
    
    return {
        "message": f"Đã xóa {deleted_count} kỳ trùng lặp",
        "deleted_count": deleted_count,
        "deleted_ids": deleted_ids,
        "kept_periods": kept_periods
    }


# ============ PRE-CLOSE VALIDATION API ============

class ValidationCheck(BaseModel):
    id: str
    name: str
    status: str  # PASS, WARN, FAIL
    severity: str  # INFO, WARNING, CRITICAL
    details: Optional[str] = None
    action_url: Optional[str] = None

class PreCloseValidationResponse(BaseModel):
    passed: bool
    checks: List[ValidationCheck]


@router.get("/periods/{period_id}/pre-close-validation", response_model=PreCloseValidationResponse)
async def get_pre_close_validation(
    period_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Get pre-close validation checks for an accounting period.
    Returns a list of automated checks before allowing period close.
    """
    # Verify period exists
    result = await db.execute(
        select(AccountingPeriodModel).where(
            AccountingPeriodModel.id == period_id,
            AccountingPeriodModel.tenant_id == tenant_id
        )
    )
    period = result.scalar_one_or_none()
    if not period:
        raise HTTPException(status_code=404, detail="Kỳ kế toán không tồn tại")
    
    checks = []
    
    # Check 1: All journals are POSTED (no DRAFT)
    draft_result = await db.execute(
        select(func.count(JournalModel.id)).where(
            JournalModel.tenant_id == tenant_id,
            JournalModel.date >= period.start_date,
            JournalModel.date <= datetime.combine(period.end_date, datetime.max.time()),
            JournalModel.status == 'DRAFT'
        )
    )
    draft_count = draft_result.scalar() or 0
    checks.append(ValidationCheck(
        id="journals_posted",
        name="Tất cả bút toán đã được duyệt",
        status="PASS" if draft_count == 0 else "FAIL",
        severity="CRITICAL",
        details=f"{draft_count} bút toán chưa duyệt (DRAFT)" if draft_count > 0 else None,
        action_url="/finance/journals?status=DRAFT" if draft_count > 0 else None
    ))
    
    # Check 2: Debit = Credit balance
    balance_result = await db.execute(
        select(
            func.coalesce(func.sum(JournalLineModel.debit), 0).label('total_debit'),
            func.coalesce(func.sum(JournalLineModel.credit), 0).label('total_credit')
        ).select_from(JournalLineModel).join(
            JournalModel, JournalLineModel.journal_id == JournalModel.id
        ).where(
            JournalModel.tenant_id == tenant_id,
            JournalModel.date >= period.start_date,
            JournalModel.date <= datetime.combine(period.end_date, datetime.max.time()),
            JournalModel.status == 'POSTED'
        )
    )
    balance = balance_result.first()
    total_debit = balance.total_debit if balance else Decimal('0')
    total_credit = balance.total_credit if balance else Decimal('0')
    is_balanced = abs(total_debit - total_credit) < Decimal('0.01')
    checks.append(ValidationCheck(
        id="balance_check",
        name="Cân đối Nợ - Có",
        status="PASS" if is_balanced else "FAIL",
        severity="CRITICAL",
        details=f"Nợ: {total_debit:,.0f} ≠ Có: {total_credit:,.0f}" if not is_balanced else f"Cân đối: {total_debit:,.0f}đ"
    ))
    
    # Check 3: No pending receivables (WARNING only)
    # NOTE: OrderModel already imported at top of file (line 27)
    # Use balance_amount > 0 to find orders with unpaid balance
    pending_ar_result = await db.execute(
        select(func.count(OrderModel.id)).where(
            OrderModel.tenant_id == tenant_id,
            OrderModel.event_date >= period.start_date,
            OrderModel.event_date <= period.end_date,
            OrderModel.balance_amount > 0  # Has unpaid balance
        )
    )
    pending_ar = pending_ar_result.scalar() or 0
    checks.append(ValidationCheck(
        id="ar_pending",
        name="Công nợ phải thu (AR)",
        status="PASS" if pending_ar == 0 else "WARN",
        severity="WARNING",
        details=f"{pending_ar} đơn hàng chưa thanh toán đầy đủ" if pending_ar > 0 else "Không có công nợ chờ",
        action_url="/finance/receivables?status=PENDING" if pending_ar > 0 else None
    ))
    
    # Check 4: Review transactions (INFO)
    txn_result = await db.execute(
        select(func.count(FinanceTransactionModel.id)).where(
            FinanceTransactionModel.tenant_id == tenant_id,
            FinanceTransactionModel.transaction_date >= period.start_date,
            FinanceTransactionModel.transaction_date <= period.end_date
        )
    )
    txn_count = txn_result.scalar() or 0
    checks.append(ValidationCheck(
        id="transactions_review",
        name="Giao dịch trong kỳ",
        status="PASS",
        severity="INFO",
        details=f"{txn_count} giao dịch đã ghi nhận"
    ))
    
    # Check 5: Bank Reconciliation Status (WARNING - manual check)
    # Check if bank reconciliation is marked as done in checklist
    bank_recon_result = await db.execute(
        select(PeriodCloseChecklistModel).where(
            PeriodCloseChecklistModel.period_id == period_id,
            PeriodCloseChecklistModel.tenant_id == tenant_id,
            PeriodCloseChecklistModel.check_name == "Đối soát ngân hàng"
        )
    )
    bank_recon_item = bank_recon_result.scalar_one_or_none()
    bank_recon_done = bank_recon_item.is_completed if bank_recon_item else False
    checks.append(ValidationCheck(
        id="bank_reconciliation",
        name="Đối soát ngân hàng",
        status="PASS" if bank_recon_done else "WARN",
        severity="WARNING",
        details="Đã hoàn tất đối soát" if bank_recon_done else "Vui lòng xác nhận đã đối soát ngân hàng",
        action_url=None if bank_recon_done else f"/finance?tab=reports"
    ))
    
    # Check 6: Pending Payables (AP) - from PurchaseOrders
    try:
        pending_ap_result = await db.execute(
            select(func.count(PurchaseOrderModel.id)).where(
                PurchaseOrderModel.tenant_id == tenant_id,
                PurchaseOrderModel.status.in_(['PENDING', 'CONFIRMED', 'RECEIVED']),
                PurchaseOrderModel.paid_amount < PurchaseOrderModel.total_amount
            )
        )
        pending_ap = pending_ap_result.scalar() or 0
    except Exception:
        pending_ap = 0  # Fallback if query fails
    
    checks.append(ValidationCheck(
        id="ap_pending",
        name="Công nợ phải trả (AP)",
        status="PASS" if pending_ap == 0 else "WARN",
        severity="WARNING",
        details=f"{pending_ap} PO chưa thanh toán đầy đủ" if pending_ap > 0 else "Không có công nợ phải trả pending",
        action_url="/finance?tab=payables" if pending_ap > 0 else None
    ))
    
    # Calculate overall pass/fail
    has_critical_fail = any(c.status == "FAIL" and c.severity == "CRITICAL" for c in checks)
    
    return PreCloseValidationResponse(
        passed=not has_critical_fail,
        checks=checks
    )


# ============ PERIOD AUDIT LOG API ============

class AuditLogEntry(BaseModel):
    id: UUID
    action: str
    performed_at: datetime
    performed_by: Optional[UUID] = None
    reason: Optional[str] = None

    class Config:
        from_attributes = True


@router.get("/periods/{period_id}/audit-log")
async def get_period_audit_log(
    period_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Get audit log entries for an accounting period.
    """
    result = await db.execute(
        select(PeriodAuditLogModel).where(
            PeriodAuditLogModel.tenant_id == tenant_id,
            PeriodAuditLogModel.period_id == period_id
        ).order_by(desc(PeriodAuditLogModel.performed_at))
    )
    logs = result.scalars().all()
    
    return [
        {
            "id": str(log.id),
            "action": log.action,
            "performed_at": log.performed_at.isoformat() if log.performed_at else None,
            "performed_by": str(log.performed_by) if log.performed_by else None,
            "reason": log.reason
        }
        for log in logs
    ]


# ============ PERIOD CLOSE CHECKLIST API ============

# Default checklist items to create for a period
DEFAULT_CHECKLIST_ITEMS = [
    {"key": "journals_posted", "name": "Duyệt tất cả bút toán", "order": 1, "automated": True},
    {"key": "bank_reconciled", "name": "Đối soát ngân hàng", "order": 2, "automated": False},
    {"key": "ar_reconciled", "name": "Đóng công nợ phải thu", "order": 3, "automated": False},
    {"key": "ap_reconciled", "name": "Đóng công nợ phải trả", "order": 4, "automated": False},
    {"key": "balance_check", "name": "Kiểm tra cân đối", "order": 5, "automated": True},
    {"key": "reports_generated", "name": "Tạo báo cáo tài chính", "order": 6, "automated": False},
    {"key": "final_approval", "name": "Phê duyệt cuối cùng", "order": 7, "automated": False},
]


@router.get("/periods/{period_id}/checklist")
async def get_period_checklist(
    period_id: UUID,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Get or initialize close checklist for an accounting period.
    """
    # Check if checklist exists
    result = await db.execute(
        select(PeriodCloseChecklistModel).where(
            PeriodCloseChecklistModel.tenant_id == tenant_id,
            PeriodCloseChecklistModel.period_id == period_id
        ).order_by(PeriodCloseChecklistModel.check_order)
    )
    items = result.scalars().all()
    
    # If no checklist, create default items
    if not items:
        for item in DEFAULT_CHECKLIST_ITEMS:
            new_item = PeriodCloseChecklistModel(
                tenant_id=tenant_id,
                period_id=period_id,
                check_key=item["key"],
                check_name=item["name"],
                check_order=item["order"],
                is_automated=item["automated"],
                is_completed=False
            )
            db.add(new_item)
        await db.commit()
        
        # Re-fetch
        result = await db.execute(
            select(PeriodCloseChecklistModel).where(
                PeriodCloseChecklistModel.tenant_id == tenant_id,
                PeriodCloseChecklistModel.period_id == period_id
            ).order_by(PeriodCloseChecklistModel.check_order)
        )
        items = result.scalars().all()
    
    completed_count = sum(1 for i in items if i.is_completed)
    total_count = len(items)
    
    return {
        "period_id": str(period_id),
        "progress": {
            "completed": completed_count,
            "total": total_count,
            "percentage": round(completed_count / total_count * 100, 1) if total_count > 0 else 0
        },
        "items": [
            {
                "id": str(item.id),
                "key": item.check_key,
                "name": item.check_name,
                "order": int(item.check_order),
                "is_automated": item.is_automated,
                "is_completed": item.is_completed,
                "completed_by": str(item.completed_by) if item.completed_by else None,
                "completed_at": item.completed_at.isoformat() if item.completed_at else None,
                "notes": item.notes
            }
            for item in items
        ]
    }


class UpdateChecklistItemRequest(BaseModel):
    is_completed: bool
    notes: Optional[str] = None


@router.patch("/periods/{period_id}/checklist/{item_id}")
async def update_checklist_item(
    period_id: UUID,
    item_id: UUID,
    request: UpdateChecklistItemRequest,
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db)
):
    """
    Update a checklist item's completion status.
    """
    result = await db.execute(
        select(PeriodCloseChecklistModel).where(
            PeriodCloseChecklistModel.id == item_id,
            PeriodCloseChecklistModel.period_id == period_id,
            PeriodCloseChecklistModel.tenant_id == tenant_id
        )
    )
    item = result.scalar_one_or_none()
    
    if not item:
        raise HTTPException(status_code=404, detail="Checklist item không tồn tại")
    
    item.is_completed = request.is_completed
    item.notes = request.notes
    if request.is_completed:
        item.completed_at = datetime.now()
        # Note: completed_by should be set from auth context
    else:
        item.completed_at = None
        item.completed_by = None
    
    await db.commit()
    await db.refresh(item)
    
    return {
        "id": str(item.id),
        "key": item.check_key,
        "name": item.check_name,
        "is_completed": item.is_completed,
        "completed_at": item.completed_at.isoformat() if item.completed_at else None,
        "notes": item.notes
    }


# ============ E1: ACTUAL COGS FROM INVENTORY (PRD Gap Fix) ============

from backend.modules.inventory.domain.models import (
    InventoryTransactionModel as InvTxnModel,
    InventoryItemModel as InvItemModel,
    InventoryLotModel as InvLotModel,
    InventoryStockModel as InvStockModel,
)


class COGSItem(BaseModel):
    """Single item COGS breakdown"""
    item_id: str
    item_name: str
    category: Optional[str]
    total_qty_used: float
    avg_unit_cost: float
    total_cost: float


class COGSReport(BaseModel):
    """COGS Report using actual inventory transaction costs"""
    period_start: date
    period_end: date
    total_cogs: float
    items_count: int
    items: List[COGSItem]


@router.get("/reports/cogs", response_model=COGSReport)
async def get_cogs_report(
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
    start_date: date = Query(None),
    end_date: date = Query(None),
):
    """
    E1: Get actual COGS from inventory EXPORT transactions.
    Uses real unit_price from inventory_transactions instead of static cost_price.
    Reference: PRD-luong-nghiep-vu-kho-hang-v2.md (GAP-FINANCE-1)
    """
    now = datetime.now()
    if not start_date:
        start_date = date(now.year, now.month, 1)
    if not end_date:
        end_date = now.date()

    # Query EXPORT transactions with actual unit_price
    result = await db.execute(
        select(
            InvTxnModel.item_id,
            InvItemModel.name.label('item_name'),
            InvItemModel.category,
            func.sum(InvTxnModel.quantity).label('total_qty'),
            func.avg(InvTxnModel.unit_price).label('avg_cost'),
            func.sum(InvTxnModel.quantity * func.coalesce(InvTxnModel.unit_price, InvItemModel.cost_price, 0)).label('total_cost')
        )
        .join(InvItemModel, InvTxnModel.item_id == InvItemModel.id)
        .where(
            InvTxnModel.tenant_id == tenant_id,
            InvTxnModel.transaction_type == 'EXPORT',
            InvTxnModel.is_reversed == False,
            func.date(InvTxnModel.created_at) >= start_date,
            func.date(InvTxnModel.created_at) <= end_date,
        )
        .group_by(InvTxnModel.item_id, InvItemModel.name, InvItemModel.category)
        .order_by(desc('total_cost'))
    )
    rows = result.all()

    items = []
    total_cogs = 0.0
    for row in rows:
        cost = float(row.total_cost or 0)
        total_cogs += cost
        items.append(COGSItem(
            item_id=str(row.item_id),
            item_name=row.item_name or "N/A",
            category=row.category,
            total_qty_used=float(row.total_qty or 0),
            avg_unit_cost=float(row.avg_cost or 0),
            total_cost=cost,
        ))

    return COGSReport(
        period_start=start_date,
        period_end=end_date,
        total_cogs=total_cogs,
        items_count=len(items),
        items=items,
    )


# ============ E2: LOT-BASED STOCK VALUATION (Balance Sheet) ============

class StockValuationItem(BaseModel):
    """Single item stock valuation"""
    item_id: str
    item_name: str
    category: Optional[str]
    uom: str
    current_stock: float
    lot_based_value: float
    static_cost_value: float
    lot_count: int


class StockValuationReport(BaseModel):
    """Stock Valuation Report using lot unit_cost"""
    valuation_date: date
    total_value_lot_based: float
    total_value_static: float
    total_items: int
    items: List[StockValuationItem]


@router.get("/reports/stock-valuation", response_model=StockValuationReport)
async def get_stock_valuation(
    tenant_id: UUID = Depends(get_current_tenant),
    db: AsyncSession = Depends(get_db),
):
    """
    E2: Get stock valuation using lot-based unit_cost.
    Compares lot-based valuation vs static cost_price valuation.
    Reference: PRD-luong-nghiep-vu-kho-hang-v2.md (E2)
    """
    # Query all items with their current stock
    items_result = await db.execute(
        select(
            InvItemModel,
            func.coalesce(func.sum(InvStockModel.quantity), 0).label('current_stock')
        )
        .outerjoin(InvStockModel, InvItemModel.id == InvStockModel.item_id)
        .where(InvItemModel.tenant_id == tenant_id)
        .group_by(InvItemModel.id)
    )
    item_rows = items_result.all()

    # Query active lots with remaining value
    lots_result = await db.execute(
        select(
            InvLotModel.item_id,
            func.sum(InvLotModel.remaining_quantity * InvLotModel.unit_cost).label('lot_value'),
            func.count(InvLotModel.id).label('lot_count'),
        )
        .where(
            InvLotModel.tenant_id == tenant_id,
            InvLotModel.status == 'ACTIVE',
            InvLotModel.remaining_quantity > 0,
        )
        .group_by(InvLotModel.item_id)
    )
    lot_map = {str(row.item_id): (float(row.lot_value or 0), int(row.lot_count or 0)) for row in lots_result.all()}

    items = []
    total_lot = 0.0
    total_static = 0.0

    for row in item_rows:
        item = row[0]
        stock = float(row.current_stock or 0)
        static_val = stock * float(item.cost_price or 0)
        lot_val, lot_count = lot_map.get(str(item.id), (0.0, 0))

        # Use lot-based value if available, otherwise fallback to static
        effective_lot_val = lot_val if lot_val > 0 else static_val

        total_lot += effective_lot_val
        total_static += static_val

        if stock > 0:
            items.append(StockValuationItem(
                item_id=str(item.id),
                item_name=item.name,
                category=item.category,
                uom=item.uom or "kg",
                current_stock=stock,
                lot_based_value=effective_lot_val,
                static_cost_value=static_val,
                lot_count=lot_count,
            ))

    # Sort by value descending
    items.sort(key=lambda x: x.lot_based_value, reverse=True)

    return StockValuationReport(
        valuation_date=datetime.now().date(),
        total_value_lot_based=total_lot,
        total_value_static=total_static,
        total_items=len(items),
        items=items,
    )
